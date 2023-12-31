# Main Code - File Update Checker
LocalVersion=1.0
# 1. Check if this code didnt updated (IDK IF REALLY NEEDED)
# 1.1 (Display message if necessary)
# 2. Check if targeted file updated
# 2.0 UI INTRODUCTION # Every step afterwards needs UI Element
# 2.1 (Update file from server if needed)
# 2.2 (Update file on server if local is newer)
# 2.3 (Be able to sync file or notify on my phone)
##############################################################################################
# Imports
import ConfigHandler as Config
import openpyxl
import logging
import os
from github import Github , GithubException
from datetime import datetime
import requests
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
##############################################################################################
# Declarations
#InputFileToRead=Config.Read('FILE','targetfile')
#InputToken=Config.Read('GITHUB','token')
#InputToken=''
###############################################################################################
# Logging #NOT USED HERE BCS THIS CODE WIIL ALWAYS RUN IN DEBUG ONLY
def Logging():
    root_logger = logging.getLogger()
    for handler in root_logger.handlers:
        root_logger.removeHandler(handler)
        handler.close()
    if not root_logger.hasHandlers():
        root_logger.setLevel(Config.Read('APP', 'logging', 'int'))
        if Config.Read('APP', 'errorlogging', 'bool'):
            file_handler = logging.FileHandler("app.log")
            file_handler.setLevel(logging.DEBUG)
            file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            file_handler.setFormatter(file_formatter)
            root_logger.addHandler(file_handler)
        if os.path.exists("app.log") and Config.Read('APP', 'errorlogging', 'bool') == False :
            os.remove("app.log")
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.DEBUG)
        console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        console_handler.setFormatter(console_formatter)
        root_logger.addHandler(console_handler)
##############################################################################################
logging.debug('Loading Main Code') # :)
##############################################################################################
# Classes and Definitions
def FileDeletion(file):
    if os.path.exists(file):
        try:
            os.remove(file)
        except:
            logging.debug('Deleting of file '+file+' failed!')
    else:
        logging.debug('Unable to locate file '+file)
##############################################################################################
class FileVersionCheck():
    Final=[]
    GitPull=[]
    VersionFinal=[]
    def LocalFileVersionExcel(FilepathToRead):
        try :
            WB = openpyxl.load_workbook(FilepathToRead,data_only=True)['Main']
        except Exception as e:
            logging.debug(str(e))
            logging.warning('Loading of local workbook '+FilepathToRead+' failed!')
            return -1
        logging.debug('Local .xlsx version: '+str(WB.cell(1,1).value)[3:])
        return float(str(str(WB.cell(1,1).value)[3:]).replace(',','.'))
    
    def LocalFileVersionOS(FilepathToRead):
        mod_time = []
        try : 
            mod_time.append(os.path.getmtime(FilepathToRead))
            mod_time.append(datetime.fromtimestamp(os.path.getmtime(FilepathToRead)).strftime('%Y-%m-%d %H:%M:%S'))
        except Exception as e:
            logging.debug(str(e))
            logging.warning('Reading of modified time failed for file '+str(FilepathToRead))
            return [int(-1),int(-1)]
        logging.debug('Local File :'+str(mod_time))
        return mod_time
    
    def GithubFileVersion(FileToRead):
        DownloadFile('version.ini')
        try :
            value1=Config.Read('APP','version','float','version.ini')
        except :
            value1=-1
        try :
            value2=Config.Read('VERSION',FileToRead,'float','version.ini')
        except:
            value2=-1
        values=[value1,value2]
        for i in range(len(values)): #IDK what is writing None on error file input so this is quick fix
            if values[i] == None:
                values[i] = -1
        FileDeletion('version.ini')
        logging.debug('Github Pull : '+str(values))
        FileVersionCheck.GitPull = values
        return values    
    
    def VersionCompareOS(Data): 
        Result = []
        for each in Data:
            if all(isinstance(value, (int, float)) for value in Data[each]):
                if Data[each].count(max(Data[each])) > 1 : Result.append(0)
                else : Result.append(Data[each].index(max(Data[each]))+1)
        else : Result.append(0) #WEIRD ERROR  
        logging.debug('Version Compare: App-'+str(Result[0])+' File-'+str(Result[1]))
        logging.debug('0=Same,1=Remote is newer,2=Local is newer')
        FileVersionCheck.VersionFinal = Result
        #return Result #calling unnecessary two times (i think return coding is not great for not changing values)
        
    def __runos__(FileToRead,FilepathToRead):
        Result = {'App':[],'File':[],'Time':[]}
        Result['App'].append(LocalVersion)
        FileVersionCheck.GithubFileVersion(FileToRead)
        Result['App'].append(FileVersionCheck.GitPull[0])
        Result['File'].append(FileVersionCheck.GitPull[1])
        Result['File'].append(FileVersionCheck.LocalFileVersionOS(FilepathToRead)[0])
        try :
            Result['Time'].append(datetime.fromtimestamp(FileVersionCheck.GitPull[1]).strftime('%Y-%m-%d %H:%M:%S'))
        except : Result['Time'].append(int(-1))
        Result['Time'].append(FileVersionCheck.LocalFileVersionOS(FilepathToRead)[1])
        logging.info('Version Checks: '+str(Result))
        FileVersionCheck.Final=Result
        FileVersionCheck.VersionCompareOS(Result)
        #return Result #Should clean all not needed return like this one
##############################################################################################
def DownloadFile(Filename,Filepath='',NewTime='',ForceDelete=0):
        if '.xlsx' in Filename:
            FileP = Filename[:Filename.rfind('.')]+'.data' #BEATUFIL REPLACING OF FILETYPE
        else : FileP = Filename
        
        try:
            if Config.Read('GITHUB','token'):
                server = Github(Config.Read('GITHUB','token'))
                repo = server.get_repo('TheAsKo/FileUpdater')
            else:
                raise ValueError("Invalid or empty GitHub token provided.")
        except GithubException as e:
            if "Bad credentials" in str(e):
                logging.warning("Invalid GitHub token provided.")
            else:
                logging.warning(f"Connecting to the server failed: {e}")
            
        try :
            Content=repo.get_contents('data/'+FileP).decoded_content
        except Exception as e :
            logging.warning(e)
            return -1
        Github.close(server)
        if ForceDelete == 1:
            FileDeletion(Filename)
        try:
            with open(Filename,'wb') as file:
                file.write(Content)    
        except: 
            logging.warning('Failed to write file')
            return -1
        if NewTime!='' and Filepath!='':
            os.utime(Filepath, (os.path.getmtime(Filepath),datetime.strptime(NewTime,'%Y-%m-%d %H:%M:%S').timestamp()))
##############################################################################################  
def UploadFile(Filename,Filepath,VersionControlledFile=True):
    if '.xlsx' in Filename:
        FileP = Filename[:Filename.rfind('.')]+'.data' #BEATUFIL REPLACING OF FILETYPE
    else : FileP = Filename
    
    try:
        if Config.Read('GITHUB','token'):
            server = Github(Config.Read('GITHUB','token'))
            repo = server.get_repo('TheAsKo/FileUpdater')
        else:
            raise ValueError("Invalid or empty GitHub token provided.")
    except GithubException as e:
        if "Bad credentials" in str(e):
            logging.warning("Invalid GitHub token provided.")
        else:
            logging.warning(f"Connecting to the server failed: {e}")
        
    with open(Filepath, 'rb') as f:
        data = f.read()
    repo.update_file('data/'+FileP,'upload excel.data', data ,requests.get('https://api.github.com/repos/TheAsKo/FileUpdater/contents/data/'+FileP).json()['sha'],branch='main')
    if VersionControlledFile == True:
        DownloadFile('version.ini')
        Config.Write('VERSION',Filename,FileVersionCheck.Final['File'][1],file='version.ini')
        with open('version.ini', 'r') as f:
            data = f.read()
        repo.update_file('data/version.ini','upload version.ini',data,requests.get('https://api.github.com/repos/TheAsKo/FileUpdater/contents/data/version.ini').json()['sha'],branch='main')
    FileDeletion('version.ini')
    Github.close(server)
##############################################################################################
def SendEmail(RecipientEmail,File,Filepath):
    try:
        Filepath=r''+Filepath
        msg = MIMEMultipart()
        msg['From'] = 'theaskobot@outlook.com'
        msg['To'] = RecipientEmail
        msg['Subject'] = 'Odoslanie suboru '+str(File)
        body = "Tato sprava bola automaticky vygenerovana , prosim neodpovedajte na nu.\nV pripade akychkolvek otazok kontakujte dodavatela softveru."
        msg.attach(MIMEText(body, 'plain'))
        with open(Filepath, 'rb') as file:  # Attach a file
            attach = MIMEApplication(file.read(), _subtype=File[File.rfind('.'):])
            attach.add_header('Content-Disposition', 'attachment', filename=File)
            msg.attach(attach)
        with smtplib.SMTP('smtp.office365.com',587) as server:
            server.starttls()  # Use TLS encryption
            server.login('theaskobot@outlook.com', 'botting123')
            server.sendmail('theaskobot@outlook.com', RecipientEmail, msg.as_string())
    except Exception as e:
        logging.warning(str(e))
        raise Exception(e)
    else:
        logging.info('Email with file '+str(File)+' was send to email '+str(RecipientEmail))


if __name__ == '__main__': #DEBUG
    logging.debug("Code Start")
    FileVersionCheck.__runos__(Config.Read('FILE','targetfile'),Config.Read('FILE','targetfilepath'))
    #DownloadFile('Zoznam.xlsx')
    #UploadFile('Zoznam.xlsx',True)
